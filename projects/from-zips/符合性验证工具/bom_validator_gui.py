import sys
import pandas as pd
from tkinter import *
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


class BOMValidator:
    def __init__(self, master):
        self.master = master
        master.title("BOM验证工具 v2.0")
        master.geometry("400x200")

        # GUI组件
        self.label = Label(master, text="选择BOM文件：")
        self.label.pack(pady=10)

        self.path_entry = Entry(master, width=40)
        self.path_entry.pack()

        self.browse_btn = Button(master, text="浏览", command=self.browse_file)
        self.browse_btn.pack(pady=5)

        self.run_btn = Button(master, text="开始验证", command=self.run_validation, state=DISABLED)
        self.run_btn.pack(pady=10)

        self.status = Label(master, text="", fg="grey")
        self.status.pack()

    def browse_file(self):
        file_path = filedialog.askopenfilename(
            filetypes=[("Excel文件", "*.xlsx"), ("All Files", "*.*")]
        )
        if file_path:
            self.path_entry.delete(0, END)
            self.path_entry.insert(0, file_path)
            self.run_btn.config(state=NORMAL)
            self.status.config(text="")

    def run_validation(self):
        input_file = self.path_entry.get()
        try:
            self.status.config(text="处理中...", fg="blue")
            self.master.update()

            # 执行验证逻辑
            self.process_file(input_file)

            messagebox.showinfo("完成", "验证报告生成成功！")
            self.status.config(text="就绪", fg="green")
        except Exception as e:
            messagebox.showerror("错误", f"处理失败：{str(e)}")
            self.status.config(text="错误发生", fg="red")

    def process_file(self, input_file):
        """核心验证逻辑"""
        # 数据读取
        ebom = pd.read_excel(input_file, sheet_name='EBOM')
        mbom = pd.read_excel(input_file, sheet_name='MBOM')

        # 数据预处理
        def preprocess(df):
            df = df.copy()
            df['件号'] = df['件号'].astype(str).str.strip()
            df['数量'] = pd.to_numeric(df['数量'], errors='coerce')
            return df.dropna(subset=['件号', '数量'])

        ebom_clean = preprocess(ebom)
        mbom_clean = preprocess(mbom)

        # 生成差异报告
        diff_report = []
        ebom_dict = ebom_clean.groupby('件号')['数量'].sum().to_dict()
        mbom_dict = mbom_clean.groupby('件号')['数量'].sum().to_dict()
        all_parts = set(ebom_dict) | set(mbom_dict)

        for part in all_parts:
            ebom_qty = ebom_dict.get(part, 0)
            mbom_qty = mbom_dict.get(part, 0)

            if part not in mbom_dict:
                diff_type = "MBOM缺失"
            elif part not in ebom_dict:
                diff_type = "EBOM缺失"
            elif ebom_qty != mbom_qty:
                diff_type = "数量不一致"
            else:
                continue

            diff_report.append({
                '差异类型': diff_type,
                '件号': part,
                'EBOM数量': ebom_qty if diff_type != "EBOM缺失" else "N/A",
                'MBOM数量': mbom_qty if diff_type != "MBOM缺失" else "N/A"
            })

        # 生成报告文件
        output_file = "BOM_对比报告.xlsx"
        with pd.ExcelWriter(output_file) as writer:
            ebom.to_excel(writer, sheet_name='EBOM原始数据', index=False)
            mbom.to_excel(writer, sheet_name='MBOM原始数据', index=False)
            pd.DataFrame(diff_report).to_excel(writer, sheet_name='差异汇总', index=False)

        # 高亮异常数据
        red_fill = PatternFill(start_color='FFC7CE', fill_type='solid')
        wb = load_workbook(output_file)

        # EBOM标记
        ebom_issues = {r['件号'] for r in diff_report if r['差异类型'] in ['MBOM缺失', '数量不一致']}
        ws_ebom = wb['EBOM原始数据']
        for row in ws_ebom.iter_rows(min_row=2):
            part = str(row[ebom.columns.get_loc('件号')].value).strip()
            if part in ebom_issues:
                for cell in row:
                    cell.fill = red_fill

        # MBOM标记
        mbom_issues = {r['件号'] for r in diff_report if r['差异类型'] in ['EBOM缺失', '数量不一致']}
        ws_mbom = wb['MBOM原始数据']
        for row in ws_mbom.iter_rows(min_row=2):
            part = str(row[mbom.columns.get_loc('件号')].value).strip()
            if part in mbom_issues:
                for cell in row:
                    cell.fill = red_fill

        # 调整列宽
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for col in ws.columns:
                max_length = max(len(str(cell.value)) for cell in col)
                ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

        wb.save(output_file)


if __name__ == "__main__":
    root = Tk()
    app = BOMValidator(root)
    root.mainloop()